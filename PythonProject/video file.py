x = 93 / 8.2572
hi_makore = 'ho hi I didnt notice you in the first time haha crazy isnt it??'
name = input("What is your name? ")
last = input("What is your last name? ")
print(f'hi {name} you bitch {last} haha')
print(hi_makore.replace('ho', 'hell no'))
mash = len(hi_makore)
rose = mash - 5
print(rose)
print('ho hi' in hi_makore)
print(hi_makore.title())
print(round(x))


nm = "DAvid"
if len(name)<3:
    print('hell no must be more than that')
if len(name)>50:
    print('its not your name.')
else:
    print(f'nice to meet you,(not really) {name.upper()}')

    wt = input('wt: ')
    matamaadif = input('l or k: ')
    if matamaadif.upper() == 'L':
        math = int(wt) / 0.45
        por = math.__round__()
        print(f'ok man do the math {por}')
    elif matamaadif.upper() == 'K':
        math = int(wt) * 0.45
        por = math.__round__()
        print(f'ok man do the math {por}')
f = 0
while f < 4:
 f += 1
 print('hell nonono')
print(f'F yeah {f} times')

es = 9
noneed = 0
ten = 3
while noneed< ten:
    noneed += 1
    shuv = (input('matachoshev?: '))
    if int(shuv) == es:
        print(f'you lucky bustered the number was {es} :( f off')
        break
else:
 print(f'I knew you will feil the num was {es} :( f off')



run = ""
nice = False

while True:
    run = input('f off ').lower()
    if run == 'h':
        print('''
no help
for you!!!!
F off you bitch!!!!''')

    elif run == 's':
        if nice:
            print('cant start twice F off!')
        else:
            nice = True
            print('Car started you moron')

    elif run == 'st':
        if not nice:
            print("already stopped")
        else:
            nice = False
            print('car stopped')

    elif run == 'q':
        break

    else:
        print('wtf is that doing??!!?')


prce = [10, 20, 30]
total = 0
for beza in prce:
    total += beza
print(f'{total}')


nun = [5,2,5,2,2]
for i in nun:
    dash = 'x'*i
    print(dash)


num = [5,2,5,2,2]
for bezasheli in num:
    tozi = ''
    for zain in range(bezasheli):
        tozi += 'x'
    print(tozi)



easy = [5,10,8,6,4,2,1]
tam = easy[0]
for b in easy:
    if b > tam:
        tam = b
print(tam)



ma = [
    [5, 2, 5],
    [2, 2, 6],
    [7, 8, 9]
]
for row in ma:
    for col in row:
      dash = 'x'
      print(dash*col)




list = [2, 2, 4, 5, 5, 6, 9, 8, 9]
a = []

for x in list:
    if x in a:
        continue
    else:
        a.append(x)

print(a)


easss = [2, 5, 14, 566]
a, b, c, d = easss
print(a, b, c, d)



s = input('enter phone number: ')
custumr = {
 '1': 'one',
 '2': 'two',
 '3': 'three',
 '4': 'four',
}
z = ""
for b in s:
    z += custumr.get(b, "!") + " "
print(z)



message = input("Enter your message: ")
words = message.split(' ')
emoji = {
    ";)": "😁",
    ":{": "😊",
}
output = ""
for word in words:
    output += emoji.get(word, word) + " "
    print(output)




def greet_user(name, last_name):
    print(f"hello {name}")
    print(f"nice name and last name! {last_name}")

print("we are done")
greet_user("Gooda", "More")
print("finish")


def square(x):
    return x*x


print(square(5))




def emoji(message):
    words = message.split(' ')
    emojis = {
        ";)": "😁",
        ":{": "😊",
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input("Enter your message: ")
print(emoji(message))



try:
    age = int(input("age: "))
    print(age)
    income = 1000
    risk = income / age
except ZeroDivisionError:
    print("I knoew your not 0 years old like wtf...")
except ValueError:
    print("You didn't enter a number. And you knew that...")




#jksnsdfnsjnfsnfkj
print("sky is blue")

#calc the number and return it squered
def squre(num4):
    return num4 * num4




class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    def move(self):
        print("move")

    def draw(self):
        print("draw")


p = Point(10, 20)
print(p.x)


class person:
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def talk(self):
     print(f"Hello! I am {self.name} and I am {self.age} years old")


p1=person("John", 18)
p1.talk()




class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    pass

class Cat(Mammal):
    def walk(self):
        print("ok")
    def eat(self):
        print("eat")

dog1 = Dog()
dog1.walk()
cat1 = Cat()
cat1.eat()



import convertors
from convertors import kgs_to_lbs
from convertors import lbs_to_kg


print(lbs_to_kg(10))




class FindMax:
    def __init__(self):
        self.num = [10, 3, 6, 2]

    def find_max(self):
        max_val = self.num[0]
        for n in self.num:
            if n > max_val:
                max_val = n
        return max_val

if __name__ == "__main__":
    p1 = FindMax()
    print(p1.find_max())



from ecommerce.shipping import calc_shipping, gregerg, fewfwef

calc_shipping()


#mudals זה מגניב אבל יש ממש הרבה, וזה הצורה שרושמים:
import random

for i in range(3):
 print(random.randint(10, 20))
#במקום "random" או "randint" אפשר לשנות למודל אחר

import random

class Dices:
 def Dice():
    members = []
    for i in range(3):
        s = random.randint(1, 6)
        b = random.randint(1, 6)
        members.append((b, s))
    leder = random.choice(members)
    print(leder)

Dices.Dice()


#זה טוב יותר: מהזה למעלה
import random

class Dices:
    def Dice(self):
        A = random.randint(1, 6)
        B = random.randint(1, 6)
        return A, B


dice = Dices()
print(dice.Dice())

#basic of mudels:"https://docs.python.org/3/library/pathlib.html#module-pathlib"

#לבדוק אם קובץ קיים → path.exists()

#לקרוא קובץ → path.read_text()

#לכתוב קובץ → path.write_text()

#לעבור על קבצים בתיקייה → for file in path.iterdir():


from pathlib import Path

path = Path()
for file in path.glob('*.py'):
    print(file)


#הכללללללללל xl עד שרשום פה אחרת....
#מאוד חשוב!!! זה על xl
import openpyxl as xl
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['גיליון1']
cell = sheet['A1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)

#עוד xl
import openpyxl as xl
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['גיליון1']
cell = sheet['A1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    print(corrected_price)

#עוד
import openpyxl as xl
from openpyxl.chart import Bar_Chart, Reference

wb = xl.load_workbook('transaction.xlsx')
sheet = wb['גיליון1']
cell = sheet['A1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('transaction.xlsx')
#עוד
import openpyxl as xl
from openpyxl.chart import BarChart, Reference, BarChart

wb = xl.load_workbook('transaction.xlsx')
sheet = wb['גיליון1']
cell = sheet['A1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
wb.save('transaction.xlsx')
#עוד




# מייבאים את הספרייה openpyxl — ספרייה שעוזרת לנו לעבוד עם קבצי אקסל (.xlsx)
import openpyxl as xl

# מתוך הספרייה הזו אנחנו מייבאים מחלקות שקשורות לתרשימים (גרפים)
from openpyxl.chart import BarChart, Reference

# --- שלב 1: פתיחת הקובץ --- #
# טוען את קובץ האקסל בשם "transaction.xlsx"
# שים לב: הקובץ חייב להיות באותה תיקייה שבה נמצא הקובץ app.py שלך
wb = xl.load_workbook('transaction.xlsx')

# ניגשים לגיליון מסוים בתוך הקובץ (כאן הוא נקרא "גיליון1")
# אם אצלך זה Sheet1 או שם אחר, שנה בהתאם
sheet = wb['גיליון1']

# --- שלב 2: מעבר על כל השורות --- #
# אנחנו מתחילים מהשורה 2 (כי שורה 1 היא לרוב כותרות)
# וממשיכים עד השורה האחרונה שבה יש נתונים
for row in range(2, sheet.max_row + 1):
    # ניגש לעמודה השלישית (C) — שם נמצא המחיר המקורי
    cell = sheet.cell(row, 3)

    # מחשבים מחיר מתוקן: 90% מהמחיר המקורי (כלומר הנחה של 10%)
    corrected_price = cell.value * 0.9

    # ניגש לעמודה הרביעית (D) — שם נכתוב את המחיר אחרי ההנחה
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# --- שלב 3: יצירת תרשים --- #
# עכשיו נגדיר טווח של התאים שבהם יש את המחירים המתוקנים (עמודה D)
values = Reference(
    sheet,
    min_row=2,  # מתחילים מהשורה השנייה
    max_row=sheet.max_row,  # עד השורה האחרונה עם נתונים
    min_col=4,  # עמודה 4 (D)
    max_col=4  # רק עמודה אחת
)

# יוצרים תרשים מסוג עמודות (Bar Chart)
chart = BarChart()

# מוסיפים לתרשים את הנתונים שציינו למעלה
chart.add_data(values)

# מציבים את התרשים על הגיליון, ליד התא E2 (כלומר קצת מימין לנתונים)
sheet.add_chart(chart, 'E2')

# --- שלב 4: שמירה --- #
# שומרים את הקובץ המעודכן
# (אם אתה לא רוצה לדרוס את הקובץ המקורי, שנה את השם כאן)
wb.save('transaction.xlsx')

# --- סיום --- #
# עכשיו אם תפתח את הקובץ "transaction.xlsx",
# תראה עמודה חדשה עם מחירים אחרי הנחה + תרשים אוטומטי לידם 🎉





#אחד אחרון - אותו דבר כמו למעלה רק בתוך פונקצייה
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['גיליון1']


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('transaction.xlsx')



#MLops
import pandas as pd
music_data = pd.read_csv("music.csv")
music_data
#MLops
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
music_data = pd.read_csv("music.csv")
x = music_data.drop(columns=['genre'])
y = music_data['genre']

model = DecisionTreeClassifier()
model.fit(x,y)
predictions = model.predict([ [21, 1], [22, 0] ])
predictions
#MLops
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score

music_data = pd.read_csv("music.csv")
x = music_data.drop(columns=['genre'])
y = music_data['genre']

x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2)

model = DecisionTreeClassifier()
model.fit(x_train, y_train)

predictions = model.predict(x_test)
score = accuracy_score(y_test, predictions)
score
#MLops

#MLops

#MLops

#MLops

#MLops

#MLops

#MLops

#MLops